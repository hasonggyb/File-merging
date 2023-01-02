import os
import time
from tkinter import filedialog as fld
from tkinter import messagebox
from tkinter import simpledialog as sd

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

# 首先弹出注意的对话框提示
messagebox.showinfo(title="注意", message="所有合并的表格格式要基本一致！！！\n并且把所有要合并的文件都存储在一个文件夹中")
# 程序运行计时开始
t1 = time.time()
# 对话框选择文件夹，并转换成python识别的文件路径格式
filenames = fld.askdirectory(title="请确定你要合并的所有文件都在该文件夹下！！").replace("/", "\\")
# 调取文件夹下的文件列表
file_lists = os.listdir(filenames)
# print(file_lists)
# 存放带路径的文件列表，开始是空列表，后续根据append追加到此类表中
data_list = []
# 弹出框选择文件表头行数
hang_row = sd.askinteger(title="表头行数", prompt="一定确定好表头一共几行，否则可能会造成数据不完整！\n请输入你的合成表的表头行数：")
# 转文本的字段，比如身份证号码在导出表中常常会显示科学技术法，在此把身份证号码在此转成文本格式
ziduan_in = sd.askstring(title="转换文本字段", prompt="请输入转换文本的字段名称(如果没有需要转换的不填)：")
# print(ziduan_in)
# 把带路径的完整文件名用append方法依次追加到data_list列表中
for file_list in file_lists:
    data_list.append(pd.read_excel(filenames + "\\" + file_list, header=hang_row - 1))
# 通过pandas中的concat方法对所有表格进行级联
data_all = pd.concat(data_list)
# 判断是否输入了转换文本的字段，如果不为空则进行转换，否则不进行转换操作
if ziduan_in != "":
    data_all[ziduan_in] = data_all[ziduan_in].apply(np.str_)
# print(data_all)
# 把转换后的的表格以excel表格的形式存放到桌面，并命名为"合成表"
data_all.to_excel(r"C:\Users\Admin\Desktop\合并表.xlsx", sheet_name="合并表", index=False)
# 把合成表的路径存放到变量filename中
filename = r"C:\Users\Admin\Desktop\合并表.xlsx"
# 调取合成表进行格式化设置
if os.path.exists(filename):
    wd = load_workbook(filename)
    ws = wd.active
    al = Alignment(horizontal="center")  # 单元格中居中
    borders_lr = Side(border_style="thin")  # 边框线
    name_border = Border(start=borders_lr, end=borders_lr, top=borders_lr, bottom=borders_lr)  # 设置边框线

    lks = []  # 存放单元格字符串长度的列表，初始化时为空表
    for i in range(1, ws.max_column + 1):
        lk = 1  # 单元个字符串长度存放变量初始化
        for j in range(1, ws.max_row + 1):
            sz = ws.cell(row=j, column=i).value  # 调取每个单元格的值
            ws.cell(row=j, column=i).alignment = al  # 每个单元格的值都居中
            ws.cell(row=j, column=i).border = name_border  # 设置可见表边框线
            if isinstance(sz, str):
                lk1 = len(sz.encode("utf-8"))  # 中文字符串的长度调取
            else:
                lk1 = len(str(sz))  # 英文字符串的长度调取
            if lk < lk1:
                lk = lk1  # 重新赋值单元格字符串长度
        lks.append(lk)  # 追加单元格字符串长度的每个值到lks列表中
    for i in range(1, ws.max_column + 1):  # 列宽自适应宽度实现循环
        k = get_column_letter(i)
        ws.column_dimensions[k].width = lks[i - 1] + 3  # 后边的3是多出3个字符空间
    wd.close()
    wd.save(r"C:\Users\Admin\Desktop\合并表.xlsx")
t2 = time.time()
timediff = t2 - t1
messagebox.showinfo(title="确定完成", message="已完成工作，共耗时：%d秒。\n合并后的表格命名为'合并表'存放在桌面上了哈！"
                                          "\n制作：九粟 \n版本：1.02（2022-12-25）" % (timediff))
